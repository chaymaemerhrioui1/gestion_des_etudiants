import io
from tkinter import *
from datetime import date
from tkinter import filedialog, simpledialog
from tkinter import messagebox
from PIL import ImageTk, Image
import os
from tkinter.ttk import Combobox
import openpyxl
from openpyxl import Workbook
import pathlib
import mysql.connector

from forgotID import forgot_my_id


filenameOpened = False

background = "#06283D"
frambg = "#EDEDED"
framfg = "#06283D"

regi = Tk()
regi.title("Registrement d Etudiant")
regi.geometry("1250x700+210+100")
regi.config(bg=background)


# exit window
def Exit():
    regi.destroy()


def registrement_page():
    regi.destroy()
    import signup


def registr():
    regi.destroy()
    import seconnecter


def get_id_tab():
    import tkinter as tk
    root = tk.Tk()
    root.geometry("200x100")

    header = tk.Label(root, text="your code is incorrect, please verify code or id ", font=("Arial", 9))
    header.pack(pady=20)

    button_frame = tk.Frame(root)
    button_frame.pack()

    ok_button = tk.Button(button_frame, text="forgot my id -_-",
                          command=lambda: [forgot_my_id(), root.destroy()])
    ok_button.pack(side=tk.LEFT, padx=10)

    cancel_button = tk.Button(button_frame, text="nah", command=lambda: [messagebox.showinfo("pressed nah",
                                                                                             "then try again"),
                                                                                              root.destroy()])
    cancel_button.pack(side=tk.LEFT, padx=10)

    root.mainloop()

# show image
def showimage():
    global filename
    global img
    filename = filedialog.askopenfilename(initialdir=os.getcwd(), title="Choisir une image", filetype=(
    ("JPG File", "*.jpg"), ("PNG File", "*.png"), ("All files", "*.txt")))
    global filenameOpened
    filenameOpened= True
    img = (Image.open(filename))
    resized_image = img.resize((190, 190))
    photo2 = ImageTk.PhotoImage(resized_image)
    lbl.config(image=photo2)
    lbl.image = photo2


# registration no.

def registration_no():
    file = openpyxl.load_workbook('Student_data.xlsx')
    sheet = file.active
    row = sheet.max_row
    max_row_value = sheet.cell(row=row, column=1).value
    try:
        Registration.set(max_row_value + 1)
    except:
        Registration.set("1")


# clear
def clear():
    global img
    Nom.set('')
    Date_naissance.set('')
    Baccalauriat.set('')
    Nationalite.set('')
    Class.set("choisir votre classe")
    f_name.set('')
    f_occupation.set('')
    m_name.set('')
    m_occupation.set('')

    registration_no()
    saveButton.config(state='normal')
    img1 = PhotoImage(file="upload photo.png")
    lbl.config(image=img1)
    lbl.image = img1
    img = ""


############save############
def verif(num):
    try:
        db = mysql.connector.connect(host='localhost', user='root', database='userdata')
        mycursor = db.cursor()
    except mysql.connector.Error:
        messagebox.showerror('Error', 'Data Connectivity Issue. Please Try Again')
        return
    query = 'select password from data where id = %s'
    mycursor.execute(query, (num,))
    row = mycursor.fetchone()

    while True:
        if row == None:
            messagebox.showerror('Error', 'invalid ID number')
            get_id_tab()

        else:
            code = simpledialog.askstring(title="data input vaerification",
                                           prompt="to assure that your are making modifications in your zone, Please enter your code:")
            print(code)
            if code is None:
                break
            elif code == str(row[0]):
                print(f'{code} and {row} match')
                return 1
            else:
                get_id_tab()



def save():
    R1 = Registration.get()
    N1 = Nom.get()
    C1 = Class.get()
    try:
        G1 = genre
    except:
        messagebox.showerror("Erreur", "Selectionner le genre !!")

    D1 = Date_naissance.get()
    Re = Baccalauriat.get()
    S1 = Nationalite.get()
    nompere = f_name.get()
    travailp = f_occupation.get()
    nommere = m_name.get()
    travailm = m_occupation.get()

    if N1 == "" or C1 == "choisir votre classe" or Re == "" or S1 == "" or nommere == "" or nompere == "" or travailp == "" or travailm == "":
        messagebox.showerror("Erreur", "Quelques données manquent !!")
    else:

        R1 = Registration.get()
        N1 = Nom.get().strip()
        nom = N1.split('_')[0]
        prenom = N1.split('_')[1]
        C1 = Class.get().strip()
        selection()
        G1 = genre
        D1 = Date_naissance.get().strip()
        Re = Baccalauriat.get()
        S1 = Nationalite.get()
        nompere = f_name.get().strip()
        travailp = f_occupation.get()
        nommere = m_name.get().strip()
        travailm = m_occupation.get().strip()
        img_variable = img
        #img_variable.show()

        # if the user wants to commit changes in his form
        if verif(R1)==1:


            try:
                db = mysql.connector.connect(host='localhost', user='root', database='userdata')
                mycursor = db.cursor()
            except mysql.connector.Error:
                messagebox.showerror('Error', 'Data Connectivity Issue. Please Try Again')
                return
            try:
                q1 = ['date_naissance date', 'genre varchar(10)', 'class varchar(15)', 'nationalite varchar(15)',
                      'bacaloreat varchar(20)', 'nom_pere varchar(50)', 'travaille_pere varchar(50)',
                      'nom_mere varchar(50)', 'travaille_mere varchar(50)', 'image longblob']

                query = 'CREATE DATABASE IF NOT EXISTS userdata'
                mycursor.execute(query)
                query = 'USE userdata'
                mycursor.execute(query)

                for i in range(len(q1)):
                    query = (f'''ALTER TABLE data
                       ADD {q1[i]} NULL''')
                    mycursor.execute(query)
            except mysql.connector.Error:
                mycursor.execute('USE userdata')

            sql = ("UPDATE data "
                   f"set  first_name='{nom}', last_name='{prenom}', date_naissance= '{D1}', genre='{G1}', class='{C1}', bacaloreat='{Re}',nationalite='{S1}', nom_pere='{nompere}', travaille_pere='{travailp}', nom_mere='{nommere}', travaille_mere= '{travailm}' "
                   f"where id = {R1}")

            mycursor.execute(sql)
            db.commit()
        else:
            import tkinter as tk
            root = tk.Tk()
            root.withdraw()  # Hide the root window

            messagebox.showinfo("Message", "Your modification wasn't committed -_-")

        #    img_data= img.read()

        if (not filenameOpened):
            import tkinter as tk
            root = tk.Tk()
            root.withdraw()  # Hide the root window
            messagebox.Message("you haven't uploaded a picture -_- however your changes have been committed.")
        # Convert binary data to MySQL binary format
        else:
            with open(filename, "rb") as f:

                imgg = Image.open(f)
                imgg_byte = imgg.tobytes()

            query = f'''UPDATE data 
                    set image=(%s)
                    where id= {R1}'''
            mycursor.execute(query, (imgg_byte,))
            db.commit()


# search """"""""""""""
def chercher():
    text = search.get()
    clear()
    saveButton.config(state='disable')

    try:
        db = mysql.connector.connect(host='localhost', user='root', database='userdata')
        mycursor = db.cursor()

    except mysql.connector.Error:
        messagebox.showerror('Error', 'Data Connectivity Issue. Please Try Again')
        return

    query = 'SELECT * FROM data WHERE id=%s'
    mycursor.execute(query, (text,))

    # Vérifier que l'utilisateur existe déjà
    row = mycursor.fetchone()
    if row is None:
        messagebox.showerror('Error', "ID doesn't exist, verify your ID")
    #    elif row>1:
    #       messagebox.showerror('Error', "there has been an error in our database.\nerror number: 69 ")
    else:

        Registration.set(row[0])
        Nom.set([row[2], row[3]])
        Class.set(row[7])
        if row[6] == 'Female':
            R2.select()
        else:
            R1.select()

        Date_naissance.set(row[5])
        Baccalauriat.set(row[8])
        Nationalite.set(row[9])
        f_name.set(row[10])
        f_occupation.set(row[11])
        m_name.set(row[12])
        m_occupation.set(row[13])

        img = Image.open(io.BytesIO(row[14]))

        resized_image = img.resize((190, 190))
        photo2 = ImageTk.PhotoImage(resized_image)
        lbl.config(image=photo2)
        lbl.image = photo2
        lbl.pack()


def update():
    import mysql.connector

    R1 = Registration.get()
    N1 = Nom.get()
    nom = N1.split('_')[0]
    prenom = N1.split('_')[1]
    C1 = Class.get()
    selection()
    G1 = genre
    D1 = Date_naissance.get()
    Re = Baccalauriat.get()
    S1 = Nationalite.get()
    nompere = f_name.get()
    travailp = f_occupation.get()
    nommere = m_name.get()
    travailm = m_occupation.get()

    try:
        db = mysql.connector.connect(host='localhost', user='root')
        mycursor = db.cursor()
    except mysql.connector.Error:
        messagebox.showerror('Error', 'Data Connectivity Issue. Please Try Again')
        return
    try:
        query = 'CREATE DATABASE IF NOT EXISTS userdata'
        mycursor.execute(query)
        query = 'USE userdata'
        mycursor.execute(query)
        query = 'ALER TABLE data' \
                'ADD date_naissance date NULL, genre varchar(10) null, class varchar(15) null, bacaloreat varchar(20) null' \
                'nationalite varchar(40) null, nom_pere varchar(50) null, travaille_pere varchar(50), nom_mere varchar(50), travaille_mere varchar(50), image longblob null '
        mycursor.execute(query)
    except mysql.connector.Error:
        mycursor.execute('USE userdata')

    sql = ("UPDATE data "
           f"set  first_name='{nom}', last_name='{prenom}', date_naissance= '{D1}', genre='{G1}', class='{C1}', bacaloreat='{Re}',nationalite='{S1}', nom_pere='{nompere}', travaille_pere='{travailp}', nom_mere='{nommere}', travaille_mere= '{travailm}' "
           f"where id = {R1}")

    mycursor.execute(sql)
    db.commit()


# gender
def selection():
    global genre
    value = radio.get()
    if value == 1:
        genre = "Male"
        print(genre)
    else:
        genre = "Female"
        print(genre)


# top frames
Label(regi, text="Email: chaymaemerhrioui@gmail.com", width=10, height=3,
      bg="#12c4c0", anchor='e').pack(side=TOP, fill=X)
Label(regi, text="Registrement d Etudiant", width=10, height=2,
      bg="#10ded9", fg="#fff", font='arial 20 bold').pack(side=TOP, fill=X)

# search box to update
search = StringVar()
Entry(regi, textvariable=search, width=15, bd=2, font="arial 20").place(x=820, y=70)
imageicon3 = PhotoImage(file="search.png")
srch = Button(regi, text="Rechercher", compound=LEFT, image=imageicon3, bg='#bdf0fc',
              font='arial 13 bold', command=chercher)
srch.place(x=1060, y=66)

imageicon4 = PhotoImage(file="Layer 4.png")
Update_button = Button(regi, image=imageicon4, bg='#10ded9', command=update)
Update_button.place(x=110, y=64)

# Registration and Date
Label(regi, text="Registration Non:", font="arial 13", fg=frambg, bg=background).place(x=30, y=150)
Label(regi, text="Date", font="arial 13", fg=frambg, bg=background).place(x=500, y=150)

Registration = IntVar()
Date = StringVar()

reg_entry = Entry(regi, textvariable=Registration, width=15, font="arial 10")
reg_entry.place(x=160, y=150)

registration_no()

today = date.today()
d1 = today.strftime("%Y/%m/%d")
date_entry = Entry(regi, textvariable=Date, width=15, font="arial 10")
date_entry.place(x=550, y=150)
Date.set(d1)

# student details

obj = LabelFrame(regi, text="Details d Etudiant ", font=20, bd=2, width=900, bg=frambg,
                 fg=framfg, height=250, relief=GROOVE)
obj.place(x=30, y=200)
Label(obj, text="Nom complet ", font="arial 13", bg=frambg, fg=framfg).place(x=30, y=50)
Label(obj, text="Date de naissance ", font="arial 13", bg=frambg, fg=framfg).place(x=30, y=100)
Label(obj, text="Genre ", font="arial 13", bg=frambg, fg=framfg).place(x=30, y=150)

Label(obj, text="Class ", font="arial 13", bg=frambg, fg=framfg).place(x=500, y=50)
Label(obj, text="Baccalauriat ", font="arial 13", bg=frambg, fg=framfg).place(x=500, y=100)
Label(obj, text="Nationalité ", font="arial 13", bg=frambg, fg=framfg).place(x=500, y=150)

Nom = StringVar()
name_entry = Entry(obj, textvariable=Nom, width=20, font="arial 10")
name_entry.insert(0,"example: Chadi_Mountassir")
name_entry.place(x=175, y=50)

Date_naissance = StringVar()
Date_naissance_entry = Entry(obj, textvariable=Date_naissance, width=20, font="arial 10")
Date_naissance_entry.insert(0,"2002/07/26")
Date_naissance_entry.place(x=175, y=100)

radio = IntVar()
R1 = Radiobutton(obj, text="Male", variable=radio, value=1, bg=frambg, fg=framfg, command=selection)
R1.place(x=150, y=150)

R2 = Radiobutton(obj, text="Female", variable=radio, value=2, bg=frambg, fg=framfg, command=selection)
R2.place(x=200, y=150)

Baccalauriat = Combobox(obj, values=['Bac Marocain ', 'Bac Etrangère'], font="Roboto 10", width=17, state="r")
Baccalauriat.place(x=630, y=100)
Baccalauriat.set("type de baccalauriat")

Nationalite = StringVar()
Nationalite_entry = Entry(obj, textvariable=Nationalite, width=20, font="arial 10")
Nationalite_entry.place(x=630, y=150)

Class = Combobox(obj, values=['cp1', 'cp2', 'g.i', 'g.c', 'g.m', 'id', 'geer', 'gee'], font="Roboto 10", width=17,
                 state="r")
Class.place(x=630, y=50)
Class.set("choisir votre classe")

# parents details

obj2 = LabelFrame(regi, text="Details des Parents ", font=20, bd=2, width=900, bg=frambg,
                  fg=framfg, height=250, relief=GROOVE)
obj2.place(x=30, y=470)

Label(obj2, text="Nom du père ", font="arial 13 ", bg=frambg, fg=framfg).place(x=30, y=50)
Label(obj2, text="travail du père ", font="arial 13 ", bg=frambg, fg=framfg).place(x=30, y=100)

f_name = StringVar()
f_entry = Entry(obj2, textvariable=f_name, width=20, font="arial 10")
f_entry.place(x=160, y=50)

f_occupation = StringVar()
fo_entry = Entry(obj2, textvariable=f_occupation, width=20, font="arial 10")
fo_entry.place(x=160, y=100)

Label(obj2, text="Nom du mère ", font="arial 13 ", bg=frambg, fg=framfg).place(x=500, y=50)
Label(obj2, text="travail du mère ", font="arial 13 ", bg=frambg, fg=framfg).place(x=500, y=100)

m_name = StringVar()
m_entry = Entry(obj2, textvariable=m_name, width=20, font="arial 10")
m_entry.place(x=630, y=50)

m_occupation = StringVar()
mo_entry = Entry(obj2, textvariable=m_occupation, width=20, font="arial 10")
mo_entry.place(x=630, y=100)

# image
f = Frame(regi, bd=3, bg='black', width=200, height=200, relief=GROOVE)
f.place(x=1000, y=150)

img = PhotoImage(file="upload photo.png")
lbl = Label(f, bg="black", image=img)
lbl.place(x=0, y=0)

# button

Button(regi, text="Télécharger", width=19, height=2, font="arial 12", bg="lightblue", command=showimage).place(x=1000,
                                                                                                               y=370)
saveButton = Button(regi, text="Enregistrer", width=19, height=2, font="arial 12", bg="lightgreen", command=save)
saveButton.place(x=1000, y=450)
Button(regi, text="Réinitialiser", width=19, height=2, font="arial 12", bg="lightpink", command=clear).place(x=1000,
                                                                                                             y=530)
Button(regi, text="Quitter", width=19, height=2, font="arial 12", bg="grey", command=Exit).place(x=1000, y=610)

loginButton = Button(regi, text='Retour', font=('Open Sans', '9', 'bold underline'), bg='white', fg='#12c4c0', bd=0,
                     cursor='hand2', activebackground='white', activeforeground='blue',
                     command=registr)
loginButton.place(x=15, y=75)

regi.mainloop()